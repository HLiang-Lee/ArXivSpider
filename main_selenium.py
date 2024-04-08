import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl
import os
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from utils import format_timestr
import requests
import logging

logging.basicConfig(filename="/path/to/ArXivSpider/logs/scrach_papers.log", level=logging.INFO)
logger = logging.getLogger(__name__)


category_dict = {
    "Computer Science": "cs",
    "Economics": "econ",
    "Electrical Engineering and Systems Science": "eess",
    "Mathematics": "math",
    "Physics": "physics",
    "Quantitative Biology": "q-bio",
    "Quantitative Finance": "q-fin",
    "Statistics": "stat"

}


arxiv_subjects = ["q-fin","stat","eess","econ","physics","math", "q-bio","cs"]

def initialize_driver(driver_path):
    """初始化 WebDriver"""
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--no_sandbox")
    options.add_argument("--disable-dev-shm-usage")
    
    return webdriver.Chrome(executable_path=driver_path, options=options)

def open_arxiv_search_page(driver, url):
    """打开 arXiv 搜索页面"""
    driver.get(url)

def is_website_accessible(url):  
    try:  
        response = requests.get(url, timeout=5)  
        if response.status_code == 200:  
            return True  
        else:  
            return False  
    except requests.RequestException:  
        return False  

def click_more_buttons(driver):
    """点击页面上的 'More' 按钮，加载更多文献"""
    try:
        more_buttons = driver.find_elements_by_xpath("//span[@class='is-size-7']")
        for button in more_buttons:
            driver.execute_script("arguments[0].click();", button)
            time.sleep(1)  # 等待加载
    except Exception as e:
        logger.info(e)

def scrape_page_content(driver, query, category):
    """从当前页面中提取文献信息"""
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    articles = soup.find_all('li', class_='arxiv-result')
    results = []
    for article in articles:
        title = article.find('p', class_='title is-5 mathjax').text.strip()
        authors = article.find('p', class_='authors').text.strip()
        authors = authors.replace("\n", "").replace("Authors:", "").replace(" ", "")
        abstract = article.find('span', class_='abstract-full').text.strip()
        abstract.replace("\n", "")
        abstract = abstract[:-13]
        arxiv_id = article.find('p', class_="list-title is-inline-block").text.strip()
        arxiv_id = arxiv_id[6:-13]
        pdf_links = "https://arxiv.org/pdf/" + arxiv_id
        # try:
        #     is_website_accessible(pdf_links)
        # except:
        #     tex_soupdf_linksrce="None"
        tex_source = "https://arxiv.org/src/" +arxiv_id
        # try:
        #     is_website_accessible(tex_source)
        # except:
        #     tex_source="None"
        # 定位 <p> 元素
        p_element = soup.find('p', class_='is-size-7')
        times=[]
        if p_element:
            # 获取所有 <span> 元素
            span_elements = p_element.find_all('span', class_='has-text-black-bis has-text-weight-semibold')
            # 获取每个 <span> 后面的文本内容
            for span in span_elements:
                # 下一个兄弟节点就是文本内容
                times.append(span.find_next_sibling(string=True).strip())
                # logger.info(time_content)
        if format_timestr(times[0]) is not None:
            submitted_time = format_timestr(times[0])
        else:
            submitted_time = times[0]
        if format_timestr(times[1]) is not None:
            original_time = format_timestr(times[1])
        else:
            original_time = times[1]
        # 查找具有特定类的span标签  
        span_tag = article.find('span', class_='has-text-grey-dark mathjax')  
        
        # 提取span标签中的文本  
        comments = span_tag.get_text(strip=True) if span_tag else "Not found"  
        

        results.append([title, authors, abstract, arxiv_id, pdf_links, tex_source, submitted_time, original_time, comments, query, category])
    return results

def save_to_excel(results, file_path):
    """将提取的文献信息保存到 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Title', 'Authors', 'Abstract', 'ArXiv ID', 'PDF Links', 'Tex Source', 'Submitted Time', 'Original Announced_Time', 'Comments', 'SubCategory', 'Category'])
    for result in results:
        ws.append(result)
    wb.save(file_path)

def scrape_and_save_page(driver, page_count, output_folder, query, searchtype, category):
    """爬取页面内容并保存为独立的 xlsx 文件"""
    # 等待文献加载
    wait = WebDriverWait(driver, 3)
    wait.until(EC.visibility_of_element_located((By.XPATH, "//ol[@class='breathe-horizontal']")))

    # 提取当前页面的文献信息
    page_results = scrape_page_content(driver, query, category)

    # 保存到 Excel 文件
    output_file = os.path.join(output_folder, f"page_{query}_{searchtype}_{category}_{page_count}.xlsx")
    save_to_excel(page_results, output_file)
    logger.info(f"Page {page_count} scraped and saved to {output_file}")

def merge_excel_files(input_folder, output_file, query, searchtype, category, total_pages):
    """合并所有独立的 xlsx 文件为一个总的 xlsx 文件"""
    all_results = []
    for filename in os.listdir(input_folder):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(input_folder, filename)
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                all_results.append(row)
            wb.close()
    # 创建总的 xlsx 文件并保存结果
    timestamp = time.strftime("%Y%m%d%H%M%S")
    output_file = f"arxiv_papers_{query}_{searchtype}_{total_pages}_{category}_{timestamp}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Title', 'Authors', 'Abstract', 'ArXiv ID', 'PDF Links', 'Tex Source', 'Submitted Time', 'Original Announced_Time', 'Comments', 'SubCategory', 'Category'])
    for result in all_results:
        ws.append(result)
    wb.save(os.path.join(input_folder, output_file))
    logger.info(f"All files merged and saved to {output_file}")

def scrape_arxiv_papers(url, driver_path, output_folder, merge_output_file, query, searchtype, category, max_pages):
    """爬取 arXiv 页面上的文献信息并保存到 Excel 文件"""
    # 初始化 WebDriver
    driver = initialize_driver(driver_path)

    try:
        # 打开 arXiv 搜索页面
        open_arxiv_search_page(driver, url)

        # 设置等待时间
        wait = WebDriverWait(driver, 10)

        # 循环翻页，最多翻页10次
        page_count = 0
        all_results = []
        while True:
            page_count += 1
            logger.info(f"Scraping page {page_count}")

            # 爬取页面内容并保存为独立的 xlsx 文件
            scrape_and_save_page(driver, page_count, output_folder, query, searchtype,category)
            # 查找下一页按钮
            try:
                # next_page_button = driver.find_element_by_xpath("//a[@rel='Next']")
                next_page_button = driver.find_element_by_class_name("pagination-next")
            except NoSuchElementException:
                # 如果找不到下一页按钮，说明已到达最后一页
                break

            # 点击下一页按钮
            driver.execute_script("arguments[0].click();", next_page_button)
            time.sleep(1)  # 等待加载

            # 如果已经到达最大页面数，停止翻页
            if page_count >= max_pages:
                break
        total_pages = page_count
        # 合并所有独立的 xlsx 文件为一个总的 xlsx 文件
        merge_excel_files(output_folder, merge_output_file, query, searchtype, category,total_pages)
        logger.info(f"All pages scraped and merged into {merge_output_file}")

    finally:
        # 关闭 WebDriver
        driver.quit()

# 示例用法
if __name__ == "__main__":

    # 打开xlsx文件
    workbook = openpyxl.load_workbook('ArXivCategory.xlsx')

    # 选择一个工作表
    sheet = workbook.active

    # 遍历整个工作表
    spider_queue = []
    for row in sheet.iter_rows(min_row=2,values_only=True):
        spider_queue.append((category_dict[row[0]], row[2]))

    logger.info(f"There are {len(spider_queue)} groups needed to scrach! ")
    counts =0
    for sq in spider_queue:
        query = sq[1]
        searchtype = 'all'
        category = sq[0]
        max_pages = 4
        counts += 1
        logger.info(f"[{counts}/{len(spider_queue)}] Search {category} {query}, max_results: {50*max_pages}")
        if category != "None":
            url = f'https://arxiv.org/search/{category}?query={query}&searchtype={searchtype}&abstracts=show&order=-announced_date_first&size=50'
        else:
            url = f'https://arxiv.org/search/?query={query}&searchtype={searchtype}&source=header'
        driver_path = 'chromedriver'
        output_folder = f"/path/to/ArXivSpider/output/pages_{query}_{searchtype}_{category}"
        if not os.path.exists(output_folder):
            os.mkdir(output_folder)
        merge_output_file = os.path.join(output_folder, f"arxiv_papers_{query}_{searchtype}_merged.xlsx")
        scrape_arxiv_papers(url, driver_path, output_folder, merge_output_file, query, searchtype, category, max_pages)
        